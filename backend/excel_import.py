"""Excel 导入 → 资金流单据（手工录入 / 权威性 4）"""
from __future__ import annotations

import json
import re
import time
import uuid
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from api_helpers import default_flows_json


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    s = str(h).strip()
    s = s.replace("（", "(").replace("）", ")")
    return s


def _field_for_header(h: Any) -> Optional[str]:
    s = _norm_header(h)
    if not s:
        return None
    sl = s.lower()
    if "万元" in s and "金额" in s:
        return "amount_wan"
    pairs = [
        ("unit", ["单位", "unit", "组织", "公司名称"]),
        ("amount", ["金额", "amount", "净额", "金额(元)"]),
        ("currency", ["币种", "currency"]),
        ("trade_date", ["交易日期", "trade_date", "日期", "业务日期", "发生日期"]),
        ("biz_code", ["业务编码", "biz_code", "业务代码", "业务编号"]),
        ("biz_name", ["业务名称", "biz_name", "业务类型"]),
        ("status", ["状态", "status"]),
        ("source_ref", ["备注", "source_ref", "说明"]),
        ("source_doc_id", ["单据号", "来源单据号", "source_doc_id", "业务单号", "外部单号"]),
        ("counterparty_account", ["对方账号", "对账号", "对方账户"]),
        ("bank_name", ["交易行名", "开户行", "银行行名"]),
        ("summary", ["摘要", "摘要说明"]),

    ]
    for field, aliases in pairs:
        for a in aliases:
            if s == a or sl == a.lower():
                return field
    return None


def _parse_amount(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip().replace(",", "").replace("，", "")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _parse_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    t = str(v).strip()
    if not t:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(t[:10], fmt).date()
        except ValueError:
            continue
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", t)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def _build_header_map(header_row: Tuple[Any, ...]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for i, cell in enumerate(header_row):
        f = _field_for_header(cell)
        if f and f not in out:
            out[f] = i
    return out


def parse_excel_workbook(content: bytes) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
    """
    返回 (数据行 dict 列表, 表头原始列表, 错误信息)。
    每行 dict: unit, amount, currency, trade_date, biz_code, biz_name, status, source_ref, source_doc_id, _rownum
    """
    errors: List[str] = []
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return [], [], [f"无法读取 Excel 文件: {e}"]
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if not first:
            return [], [], ["工作表为空"]
        header = tuple(first)
        hmap = _build_header_map(header)
        if "unit" not in hmap or ("amount" not in hmap and "amount_wan" not in hmap):
            return (
                [],
                list(str(x) if x is not None else "" for x in header),
                [
                    "表头需包含「单位」列以及「金额」或「金额(万元)」列。"
                    "可选：币种、交易日期、业务编码、业务名称、状态、备注、单据号。",
                ],
            )
        raw_headers = list(str(x) if x is not None else "" for x in header)
        data_rows: List[Dict[str, Any]] = []
        rownum = 1
        for row in rows_iter:
            rownum += 1
            if not row:
                continue
            def col(name: str) -> Any:
                idx = hmap.get(name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            unit = col("unit")
            amt = None
            if "amount_wan" in hmap:
                amt = _parse_amount(col("amount_wan"))
                if amt is not None:
                    amt = amt * 10000.0
            if amt is None and "amount" in hmap:
                amt = _parse_amount(col("amount"))
            unit_s = str(unit).strip() if unit is not None and str(unit).strip() else ""
            if not unit_s and amt is None:
                continue
            if not unit_s:
                errors.append(f"第{rownum}行：缺少单位")
                continue
            if amt is None:
                errors.append(f"第{rownum}行：金额无效或为空")
                continue
            cur = col("currency")
            currency = str(cur).strip().upper() if cur not in (None, "") else "CNY"
            td = _parse_date(col("trade_date")) or date.today()
            bc = col("biz_code")
            bn = col("biz_name")
            st = col("status")
            status = str(st).strip() if st not in (None, "") else "待确认"
            if status not in ("预测", "未确认", "待确认", "流水", "已确认", "待审核"):
                status = "待确认"
            ref = col("source_ref")
            source_ref = str(ref).strip()[:60] if ref not in (None, "") else None
            sdoc = col("source_doc_id")
            source_doc_id = str(sdoc).strip()[:80] if sdoc not in (None, "") else None
            biz_code = str(bc).strip() if bc not in (None, "") else None
            biz_name = str(bn).strip() if bn not in (None, "") else None
            cp = col("counterparty_account")
            bnk = col("bank_name")
            sm = col("summary")
            counterparty_account = str(cp).strip()[:100] if cp not in (None, "") else None
            bank_name = str(bnk).strip()[:200] if bnk not in (None, "") else None
            summary = str(sm).strip()[:500] if sm not in (None, "") else None
            data_rows.append(
                {
                    "unit": unit_s,
                    "amount": float(amt),
                    "currency": currency,
                    "trade_date": td,
                    "biz_code": biz_code,
                    "biz_name": biz_name,
                    "status": status,
                    "source_ref": source_ref,
                    "source_doc_id": source_doc_id,
                    "counterparty_account": counterparty_account,
                    "bank_name": bank_name,
                    "summary": summary,
                    "_rownum": rownum,
                }
            )
        return data_rows, raw_headers, errors
    finally:
        wb.close()


def run_excel_import(db, content: bytes, force_override: bool, append_sync_log) -> dict:
    """解析 Excel 并写入 CashflowRecord；append_sync_log 与 main._append_sync_log 签名一致。"""
    from database import CashflowRecord, CashflowCollection, CashflowBusiness, BizFlowInfo, CashflowSubject

    rows, headers, parse_errors = parse_excel_workbook(content)
    if parse_errors and not rows:
        return {
            "ok": False,
            "records_created": 0,
            "records_skipped": 0,
            "errors": parse_errors,
            "headers": headers,
        }

    run_id = f"excel-{date.today().strftime('%Y%m%d')}-{uuid.uuid4().hex[:10]}"
    col_code = f"BATCH{int(time.time() * 1000)}"
    col = CashflowCollection(code=col_code, source_system="手工录入")
    db.add(col)
    db.flush()

    businesses = db.query(CashflowBusiness).filter(CashflowBusiness.is_deleted == False).all()
    by_code = {b.code: b for b in businesses}
    by_name = {b.name: b for b in businesses}

    rc = db.query(CashflowRecord).count()
    created = 0
    skipped = 0
    row_errors: List[str] = list(parse_errors)
    src = "手工录入"
    rnk = 4

    for rec in rows:
        rownum = rec["_rownum"]
        bid = None
        if rec.get("biz_code") and rec["biz_code"] in by_code:
            bid = by_code[rec["biz_code"]].id
        elif rec.get("biz_name"):
            bn = rec["biz_name"]
            if bn in by_name:
                bid = by_name[bn].id
            else:
                for b in businesses:
                    if bn in b.name or b.name in bn:
                        bid = b.id
                        break

        td = rec["trade_date"]
        amt = float(rec["amount"])
        currency = rec["currency"] or "CNY"
        fj = default_flows_json(
            db, CashflowBusiness, BizFlowInfo, CashflowSubject,
            bid, amt, currency, td, rec["status"],
        )

        source_doc_id = rec.get("source_doc_id")
        if not source_doc_id:
            source_doc_id = f"{run_id}|r{rownum}"

        existing = (
            db.query(CashflowRecord)
            .filter(
                CashflowRecord.source_doc_id == source_doc_id,
                CashflowRecord.is_deleted == False,
            )
            .first()
        )
        if existing:
            if not force_override:
                skipped += 1
                append_sync_log(
                    db, run_id, "skip", src,
                    f"Excel 第{rownum}行：单据号已存在，跳过（可勾选强制覆盖后重导）",
                    target_record_id=existing.id,
                    source_doc_id=source_doc_id,
                )
                continue
            snap = json.dumps(
                {
                    "excel_row": rownum,
                    "before_amount": existing.amount,
                    "before_unit": existing.unit,
                },
                ensure_ascii=False,
            )
            existing.unit = rec["unit"]
            existing.biz_id = bid
            existing.currency = currency
            existing.amount = amt
            existing.amount_cny = amt if currency == "CNY" else existing.amount_cny
            existing.trade_date = td
            existing.settle_date = td
            existing.source_system = src
            existing.authority_rank = rnk
            existing.source_ref = rec.get("source_ref")
            existing.status = rec["status"]
            existing.flows_json = fj
            existing.counterparty_account = rec.get("counterparty_account")
            existing.bank_name = rec.get("bank_name")
            existing.summary = rec.get("summary")
            existing.collection_id = col.id
            db.flush()
            created += 1
            append_sync_log(
                db, run_id, "override", src,
                f"Excel 第{rownum}行：强制覆盖",
                snapshot_json=snap,
                target_record_id=existing.id,
                source_doc_id=source_doc_id,
            )
            continue

        rc += 1
        code = f"CF{date.today().strftime('%Y%m%d')}{rc:08d}"
        acny = amt if currency == "CNY" else None
        obj = CashflowRecord(
            code=code,
            collection_id=col.id,
            unit=rec["unit"],
            biz_id=bid,
            currency=currency,
            amount=amt,
            trade_date=td,
            settle_date=td,
            source_system=src,
            source_ref=rec.get("source_ref"),
            source_doc_id=source_doc_id,
            authority_rank=rnk,
            amount_cny=acny,
            status=rec["status"],
            flows_json=fj,
            counterparty_account=rec.get("counterparty_account"),
            bank_name=rec.get("bank_name"),
            summary=rec.get("summary"),
            is_deleted=False,
        )
        db.add(obj)
        db.flush()
        created += 1
        append_sync_log(
            db, run_id, "insert", src,
            f"Excel 导入第{rownum}行",
            target_record_id=obj.id,
            source_doc_id=source_doc_id,
        )

    db.commit()
    return {
        "ok": True,
        "run_id": run_id,
        "collection_id": col.id,
        "collection_code": col_code,
        "records_created": created,
        "records_skipped": skipped,
        "errors": row_errors,
        "headers": headers,
        "source_system": src,
    }
